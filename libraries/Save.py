

from scipy import interpolate
import json
import numpy as np
import wx
import io
import os
import pandas as pd
import openpyxl
from libraries.ConfigFile import add_core_level_Data
from openpyxl.drawing.image import Image
from openpyxl.styles import Border, Side, PatternFill, Font
from openpyxl import load_workbook
from libraries.Sheet_Operations import on_sheet_selected
from copy import deepcopy
# from Functions import convert_to_serializable_and_round


def save_all_sheets_with_plots(window):
    if 'FilePath' not in window.Data or not window.Data['FilePath']:
        wx.MessageBox("No file selected. Please open a file first.", "Error", wx.OK | wx.ICON_ERROR)
        return

    file_path = window.Data['FilePath']

    try:
        for sheet_name in window.Data['Core levels'].keys():
            # Select the sheet
            window.sheet_combobox.SetValue(sheet_name)
            from libraries.Sheet_Operations import on_sheet_selected
            on_sheet_selected(window, sheet_name)

            # Get fitting data
            fit_data = window.get_data_for_save()

            # Save fitting data to Excel
            save_to_excel(window, fit_data, file_path, sheet_name)

            # Save plot to Excel
            save_plot_to_excel(window)

        # Save results table
        save_results_table(window)

        json_file_path = os.path.splitext(file_path)[0] + '.json'
        json_data = convert_to_serializable_and_round(window.Data)
        with open(json_file_path, 'w') as json_file:
            json.dump(json_data, json_file, indent=2)

        window.show_popup_message2("Save Complete", "All sheets, plots, and results table have been saved.")

    except Exception as e:
        wx.MessageBox(f"Error saving sheets with plots: {str(e)}", "Error", wx.OK | wx.ICON_ERROR)
def save_data(window, data):
    if 'FilePath' not in window.Data or not window.Data['FilePath']:
        wx.MessageBox("No file path found in window.Data. Please open a file first.", "Error", wx.OK | wx.ICON_ERROR)
        return

    file_path = window.Data['FilePath']
    sheet_name = window.sheet_combobox.GetValue()

    try:
        # Save to Excel
        save_to_excel(window, data, file_path, sheet_name)

        # Save JSON file with entire window.Data
        json_file_path = os.path.splitext(file_path)[0] + '.json'

        # Create a copy of window.Data to modify
        json_data = window.Data.copy()

        # Convert numpy arrays and other non-serializable types to lists, and round floats
        json_data = convert_to_serializable_and_round(json_data)

        with open(json_file_path, 'w') as json_file:
            json.dump(json_data, json_file, indent=2)

        print("Data Saved")
    except Exception as e:
        wx.MessageBox(f"Error saving data: {str(e)}", "Error", wx.OK | wx.ICON_ERROR)


def convert_to_serializable_and_round(obj, decimal_places=2):
    try:
        if isinstance(obj, (float, np.float32, np.float64)):
            return round(float(obj), decimal_places)
        elif isinstance(obj, (int, np.int32, np.int64)):
            return int(obj)
        elif isinstance(obj, np.ndarray):
            return [convert_to_serializable_and_round(item, decimal_places) for item in obj.tolist()]
        elif isinstance(obj, list):
            return [convert_to_serializable_and_round(item, decimal_places) for item in obj]
        elif isinstance(obj, dict):
            return {k: convert_to_serializable_and_round(v, decimal_places) for k, v in obj.items()}
        elif isinstance(obj, wx.grid.Grid):
            return {
                "rows": obj.GetNumberRows(),
                "cols": obj.GetNumberCols(),
                "data": [[convert_to_serializable_and_round(obj.GetCellValue(row, col), decimal_places)
                          for col in range(obj.GetNumberCols())]
                         for row in range(obj.GetNumberRows())]
            }
        elif hasattr(obj, 'tolist'):  # This catches any object with a tolist method, like some lmfit results
            return convert_to_serializable_and_round(obj.tolist(), decimal_places)
        else:
            return obj
    except Exception as e:
        return str(obj)  # Return a string representation as a fallback

def convert_to_serializable(obj):
    if isinstance(obj, np.ndarray):
        return obj.tolist()
    elif isinstance(obj, dict):
        return {k: convert_to_serializable(v) for k, v in obj.items()}
    elif isinstance(obj, list):
        return [convert_to_serializable(v) for v in obj]
    elif isinstance(obj, (np.int64, np.int32, np.float64, np.float32)):
        return obj.item()
    elif isinstance(obj, wx.grid.Grid):
        return {
            "rows": obj.GetNumberRows(),
            "cols": obj.GetNumberCols(),
            "data": [[obj.GetCellValue(row, col) for col in range(obj.GetNumberCols())]
                     for row in range(obj.GetNumberRows())]
        }
    else:
        return obj


def save_to_excel(window, data, file_path, sheet_name):
    existing_df = pd.read_excel(file_path, sheet_name=sheet_name)

    # Remove previously fitted data if it exists
    if existing_df.shape[1] > 3:  # If there are more than 3 columns (BE, Raw Data, empty)
        existing_df = existing_df.iloc[:, :3]  # Keep only the first 3 columns

    # Ensure there's an empty column C
    if existing_df.shape[1] < 3:
        existing_df.insert(2, '', np.nan)

    if 'x_values' in data and data['x_values'] is not None:
        x_values = data['x_values'].to_numpy() if isinstance(data['x_values'], pd.Series) else data['x_values']
        filtered_data = pd.DataFrame({
            'BE': x_values
        })

        if data['background'] is not None and data['calculated_fit'] is not None:
            mask = np.isin(data['x_values'], x_values)
            y_values = data['y_values'][mask]

            if len(y_values) == len(data['calculated_fit']):
                residuals = y_values - data['calculated_fit']
                filtered_data['Residuals'] = residuals
            else:
                filtered_data['Residuals'] = np.nan
        else:
            filtered_data['Residuals'] = np.nan

        filtered_data['Background'] = data['background'] if data['background'] is not None else np.nan
        filtered_data['Calculated Fit'] = data['calculated_fit'] if data['calculated_fit'] is not None else np.nan

        if data['individual_peak_fits']:
            num_rows = len(x_values)
            num_peaks = data['peak_params_grid'].GetNumberRows() // 2
            for i in range(num_peaks):
                row = i * 2
                peak_label = data['peak_params_grid'].GetCellValue(row, 1)  # Get the peak label
                if i < len(data['individual_peak_fits']):
                    reversed_peak = np.array(data['individual_peak_fits'][i])[::-1]
                    trimmed_peak = reversed_peak[:num_rows]
                    filtered_data[peak_label] = trimmed_peak  # Use peak label as column name

        # Rename columns to avoid conflicts before inserting them
        for i, col in enumerate(filtered_data.columns):
            new_col_name = col
            while new_col_name in existing_df.columns:
                new_col_name += '_new'
            existing_df.insert(3 + i, new_col_name, filtered_data[col])

    # Ensure there are at least 23 columns (A to W)
    while existing_df.shape[1] < 23:
        existing_df[f'Column_{existing_df.shape[1] + 1}'] = ''

    # Rename columns starting with "Unnamed" or "Column" to empty string
    existing_df.columns = ['' if col.startswith(('Unnamed', 'Column')) else col for col in existing_df.columns]

    # Ensure column C is empty
    if existing_df.columns[2] != '':
        existing_df.rename(columns={existing_df.columns[2]: ''}, inplace=True)

    # Create DataFrame for peak fitting parameters
    peak_params_df = pd.DataFrame()
    for col in range(window.peak_params_grid.GetNumberCols()):
        col_name = window.peak_params_grid.GetColLabelValue(col)
        col_data = [window.peak_params_grid.GetCellValue(row, col) for row in
                    range(window.peak_params_grid.GetNumberRows())]
        peak_params_df[col_name] = col_data

    # Add peak_params_df to existing_df starting from column 23 (X)
    for i, col in enumerate(peak_params_df.columns):
        existing_df.insert(23 + i, col, peak_params_df[col])


    with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        existing_df.to_excel(writer, sheet_name=sheet_name, index=False)

        # Remove border from first row
        workbook = writer.book
        worksheet = workbook[sheet_name]
        for cell in worksheet[1]:
            cell.border = openpyxl.styles.Border(
                left=openpyxl.styles.Side(style=None),
                right=openpyxl.styles.Side(style=None),
                top=openpyxl.styles.Side(style=None),
                bottom=openpyxl.styles.Side(style=None)
            )

            # Define styles
            thin_side = Side(style='thin')
            thick_side = Side(style='medium')
            green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # Light green
            bold_font = Font(bold=True)

            start_row = 2  # Assuming data starts from the second row
            num_peak_rows = window.peak_params_grid.GetNumberRows()
            end_row = start_row + num_peak_rows - 1
            start_col = 24  # Column X (24th column)
            end_col = worksheet.max_column

            for row in range(start_row - 1, end_row + 1):  # Start from header row
                for col in range(start_col, end_col + 1):
                    cell = worksheet.cell(row=row, column=col)

                    # Default to thin borders
                    left = right = top = bottom = thin_side

                    # Header row
                    if row == start_row - 1:
                        cell.fill = green_fill
                        cell.font = bold_font
                        top = thick_side

                    # Add thick borders for outer edges
                    if row == start_row - 1 or row == end_row:
                        bottom = thick_side
                    if col == start_col:
                        left = thick_side
                    if col == end_col:
                        right = thick_side

                    # Add thick bottom border for every second row (constraints row)
                    if (row - start_row + 1) % 2 == 0:
                        bottom = thick_side

                    cell.border = Border(left=left, right=right, top=top, bottom=bottom)

    # After saving to Excel, update the plot with the current limits
    if hasattr(window, 'plot_config'):
        limits = window.plot_config.get_plot_limits(window, sheet_name)
        window.ax.set_xlim(limits['Xmax'], limits['Xmin'])  # Reverse X-axis
        window.ax.set_ylim(limits['Ymin'], limits['Ymax'])
        window.canvas.draw_idle()


def refresh_sheets(window, on_sheet_selected_func):
    if 'FilePath' not in window.Data or not window.Data['FilePath']:
        wx.MessageBox("No file currently open. Please open a file first.", "Error", wx.OK | wx.ICON_ERROR)
        return

    current_sheet = window.sheet_combobox.GetValue()
    file_path = window.Data['FilePath']

    try:
        # Save current state to JSON
        json_file_path = os.path.splitext(file_path)[0] + '.json'
        json_data = convert_to_serializable_and_round(window.Data)
        with open(json_file_path, 'w') as json_file:
            json.dump(json_data, json_file, indent=2)

        # Reopen the XLSX file
        excel_file = pd.ExcelFile(file_path)
        sheet_names = excel_file.sheet_names

        # Update sheet names in the combobox
        window.sheet_combobox.Clear()
        window.sheet_combobox.AppendItems(sheet_names)

        # Update window.Data with new sheet information
        for sheet_name in sheet_names:
            if sheet_name not in window.Data['Core levels']:
                window.Data = add_core_level_Data(window.Data, window, file_path, sheet_name)

        # Remove any sheets from window.Data that no longer exist in the Excel file
        sheets_to_remove = set(window.Data['Core levels'].keys()) - set(sheet_names)
        for sheet_name in sheets_to_remove:
            del window.Data['Core levels'][sheet_name]

        # Update the number of core levels
        window.Data['Number of Core levels'] = len(sheet_names)

        # Set the current sheet as selected if it still exists, otherwise select the first sheet
        if current_sheet in sheet_names:
            window.sheet_combobox.SetValue(current_sheet)
        elif sheet_names:
            window.sheet_combobox.SetValue(sheet_names[0])
            current_sheet = sheet_names[0]

        # Update the plot for the current sheet
        event = wx.CommandEvent(wx.EVT_COMBOBOX.typeId)
        event.SetString(current_sheet)
        on_sheet_selected_func(window, event)

        # Update plot limits
        window.plot_config.update_plot_limits(window, current_sheet)

        # Refresh the plot
        window.plot_manager.plot_data(window)
        window.clear_and_replot()

        wx.MessageBox(f"Sheets refreshed. Total sheets: {len(sheet_names)}", "Success", wx.OK | wx.ICON_INFORMATION)

    except Exception as e:
        import traceback
        traceback.print_exc()
        wx.MessageBox(f"Error refreshing sheets: {str(e)}", "Error", wx.OK | wx.ICON_ERROR)


def save_to_json(window, file_path):
    json_file_path = os.path.splitext(file_path)[0] + '.json'

    # Create a copy of window.Data to modify
    data_to_save = window.Data.copy()

    def convert_to_list(item):
        if isinstance(item, np.ndarray):
            return item.tolist()
        elif isinstance(item, list):
            return item
        else:
            return item  # Return as is if it's neither numpy array nor list

    # Convert numpy arrays to lists for JSON serialization
    for sheet_name, sheet_data in data_to_save['Core levels'].items():
        sheet_data['B.E.'] = convert_to_list(sheet_data['B.E.'])
        sheet_data['Raw Data'] = convert_to_list(sheet_data['Raw Data'])
        sheet_data['Background']['Bkg Y'] = convert_to_list(sheet_data['Background']['Bkg Y'])

        if 'Fitting' in sheet_data and 'Peaks' in sheet_data['Fitting']:
            for peak_label, peak_data in sheet_data['Fitting']['Peaks'].items():
                if 'Y' in peak_data:
                    peak_data['Y'] = convert_to_list(peak_data['Y'])

    with open(json_file_path, 'w') as json_file:
        json.dump(data_to_save, json_file, indent=2)

def save_data_ORI(window, data):
    if 'FilePath' not in window.Data or not window.Data['FilePath']:
        wx.MessageBox("No file path found in window.Data. Please open a file first.", "Error", wx.OK | wx.ICON_ERROR)
        return

    file_path = window.Data['FilePath']
    sheet_name = window.sheet_combobox.GetValue()

    try:
        existing_df = pd.read_excel(file_path, sheet_name=sheet_name)

        # Remove previously fitted data if it exists
        if existing_df.shape[1] > 3:  # If there are more than 3 columns (BE, Raw Data, empty)
            existing_df = existing_df.iloc[:, :3]  # Keep only the first 3 columns

        # Ensure there's an empty column C
        if existing_df.shape[1] < 3:
            existing_df.insert(2, '', np.nan)

        if 'x_values' in data and data['x_values'] is not None:
            x_values = data['x_values'].to_numpy() if isinstance(data['x_values'], pd.Series) else data['x_values']
            filtered_data = pd.DataFrame({
                'BE': x_values
            })

            if data['background'] is not None and data['calculated_fit'] is not None:
                mask = np.isin(data['x_values'], x_values)
                y_values = data['y_values'][mask]

                if len(y_values) == len(data['calculated_fit']):
                    residuals = y_values - data['calculated_fit']
                    filtered_data['Residuals'] = residuals
                else:
                    filtered_data['Residuals'] = np.nan
            else:
                filtered_data['Residuals'] = np.nan

            filtered_data['Background'] = data['background'] if data['background'] is not None else np.nan
            filtered_data['Calculated Fit'] = data['calculated_fit'] if data['calculated_fit'] is not None else np.nan

            if data['individual_peak_fits']:
                num_rows = len(x_values)
                num_peaks = data['peak_params_grid'].GetNumberRows() // 2
                for i in range(num_peaks):
                    row = i * 2
                    peak_label = data['peak_params_grid'].GetCellValue(row, 1)  # Get the peak label
                    if i < len(data['individual_peak_fits']):
                        reversed_peak = np.array(data['individual_peak_fits'][i])[::-1]
                        trimmed_peak = reversed_peak[:num_rows]
                        filtered_data[peak_label] = trimmed_peak  # Use peak label as column name

            # Rename columns to avoid conflicts before inserting them
            for i, col in enumerate(filtered_data.columns):
                new_col_name = col
                while new_col_name in existing_df.columns:
                    new_col_name += '_new'
                existing_df.insert(3 + i, new_col_name, filtered_data[col])

        # Pad with empty columns if necessary
        while existing_df.shape[1] < 3 + window.num_fitted_columns:
            existing_df[f'Empty_{existing_df.shape[1]}'] = np.nan

        # Remove names from column C and empty columns
        existing_df.columns = ['' if i == 2 or i >= 3 + len(filtered_data.columns) else col for i, col in
                               enumerate(existing_df.columns)]

        with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            existing_df.to_excel(writer, sheet_name=sheet_name, index=False)

            # Remove border from first row
            workbook = writer.book
            worksheet = workbook[sheet_name]
            for cell in worksheet[1]:
                cell.border = openpyxl.styles.Border(
                    left=openpyxl.styles.Side(style=None),
                    right=openpyxl.styles.Side(style=None),
                    top=openpyxl.styles.Side(style=None),
                    bottom=openpyxl.styles.Side(style=None)
                )

        txt_file_path = os.path.splitext(file_path)[0] + '.kfit'
        with open(txt_file_path, 'w') as txt_file:
            txt_file.write(f"Peak Fitting Properties for {os.path.basename(file_path)}\n")
            txt_file.write(f"Sheet: {sheet_name}\n\n")

            # Save background data
            if 'x_values' in data and data['x_values'] is not None and data['background'] is not None:
                txt_file.write("Background Data:\n")
                txt_file.write("X: ")
                txt_file.write(", ".join(map(str, data['x_values'])))
                txt_file.write("\n")
                txt_file.write("Y: ")
                txt_file.write(", ".join(map(str, data['background'])))
                txt_file.write("\n\n")

            if data['peak_params_grid'] is not None:
                txt_file.write("Peaks Parameters:\n")
                num_peaks = data['peak_params_grid'].GetNumberRows() // 2
                for i in range(num_peaks):
                    row = i * 2
                    peak_name = data['peak_params_grid'].GetCellValue(row, 1)
                    position = data['peak_params_grid'].GetCellValue(row, 2)
                    height = data['peak_params_grid'].GetCellValue(row, 3)
                    fwhm = data['peak_params_grid'].GetCellValue(row, 4)
                    lg_ratio = data['peak_params_grid'].GetCellValue(row, 5)

                    # Get constraints
                    position_constraint = data['peak_params_grid'].GetCellValue(row + 1, 2)
                    height_constraint = data['peak_params_grid'].GetCellValue(row + 1, 3)
                    fwhm_constraint = data['peak_params_grid'].GetCellValue(row + 1, 4)
                    lg_ratio_constraint = data['peak_params_grid'].GetCellValue(row + 1, 5)

                    txt_file.write(f"Peak {i + 1} ({peak_name}):\n")
                    txt_file.write(f"  Position: {position}\n")
                    txt_file.write(f"  Position Constraint: {position_constraint}\n")
                    txt_file.write(f"  Height: {height}\n")
                    txt_file.write(f"  Height Constraint: {height_constraint}\n")
                    txt_file.write(f"  FWHM: {fwhm}\n")
                    txt_file.write(f"  FWHM Constraint: {fwhm_constraint}\n")
                    txt_file.write(f"  L/G Ratio: {lg_ratio}\n")
                    txt_file.write(f"  L/G Ratio Constraint: {lg_ratio_constraint}\n")

            if data['results_grid'] is not None:
                txt_file.write("\nResults:\n")
                # Get all column labels dynamically
                num_cols = data['results_grid'].GetNumberCols()
                labels = [data['results_grid'].GetColLabelValue(col) for col in range(num_cols)]

                for row in range(data['results_grid'].GetNumberRows()):
                    txt_file.write(f"Row {chr(65 + row)}:\n")  # A, B, C, etc.
                    for col, label in enumerate(labels):
                        if label != '':  # Skip empty column labels
                            value = data['results_grid'].GetCellValue(row, col)
                            txt_file.write(f"  {label}: {value}\n")
                txt_file.write("\n")  # Add a blank line between peaks



        print("Data Saved")
    except Exception as e:
        import traceback
        traceback.print_exc()
        wx.MessageBox(f"Error saving data: {str(e)}", "Error", wx.OK | wx.ICON_ERROR)


def save_plot_to_excel(window):
    if 'FilePath' not in window.Data or not window.Data['FilePath']:
        wx.MessageBox("No file selected. Please open a file first.", "Error", wx.OK | wx.ICON_ERROR)
        return

    file_path = window.Data['FilePath']
    sheet_name = window.sheet_combobox.GetValue()

    try:
        # Save the current figure to a bytes buffer
        buf = io.BytesIO()
        window.figure.savefig(buf, format='png', dpi=80, bbox_inches='tight')
        buf.seek(0)

        # Open the workbook and select the sheet
        wb = openpyxl.load_workbook(file_path)

        # Check if the sheet exists, if not create it
        if sheet_name not in wb.sheetnames:
            ws = wb.create_sheet(sheet_name)
        else:
            ws = wb[sheet_name]

        # Remove existing images
        for img in ws._images:
            ws._images.remove(img)

        # Create an image from the bytes buffer
        img = Image(buf)

        # Add the image to the worksheet
        ws.add_image(img, 'A4')  # You can adjust the cell reference as needed

        # Save the workbook
        wb.save(file_path)

        # Save as PNG
        png_filename = f"{os.path.splitext(os.path.basename(file_path))[0]}_{sheet_name}.png"
        png_filepath = os.path.join(os.path.dirname(file_path), png_filename)
        window.figure.savefig(png_filepath, format='png', dpi=300, bbox_inches='tight')

        print(f"Plot saved as PNG: {png_filepath}")

        print(f"Plot saved to Excel file: {file_path}, Sheet: {sheet_name}")
        # wx.MessageBox(f"Plot saved to Excel file:\n{file_path}\nSheet: {sheet_name}", "Success",
        #               wx.OK | wx.ICON_INFORMATION)
        window.show_popup_message2("Plot saved into Excel file", f"Under sheet: {sheet_name}")

    except Exception as e:
        import traceback
        traceback.print_exc()
        wx.MessageBox(f"Error saving plot to Excel: {str(e)}", "Error", wx.OK | wx.ICON_ERROR)


from openpyxl.styles import Font, Border, Side, PatternFill, Alignment

from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
import openpyxl
import wx

from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
import openpyxl
import wx


def save_results_table(window):
    if 'FilePath' not in window.Data or not window.Data['FilePath']:
        wx.MessageBox("No file selected. Please open a file first.", "Error", wx.OK | wx.ICON_ERROR)
        return

    file_path = window.Data['FilePath']

    try:
        wb = openpyxl.load_workbook(file_path)

        sheet_name = 'Results Table'
        if sheet_name in wb.sheetnames:
            wb.remove(wb[sheet_name])
        ws = wb.create_sheet(sheet_name)

        headers = [window.results_grid.GetColLabelValue(col) for col in range(window.results_grid.GetNumberCols())]
        for col, header in enumerate(headers, start=2):
            cell = ws.cell(row=2, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for row in range(window.results_grid.GetNumberRows()):
            for col in range(window.results_grid.GetNumberCols()):
                cell = ws.cell(row=row + 3, column=col + 2, value=window.results_grid.GetCellValue(row, col))
                cell.alignment = Alignment(horizontal="center", vertical="center")

        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                             bottom=Side(style='thin'))
        thick_border = Border(left=Side(style='medium'), right=Side(style='medium'), top=Side(style='medium'),
                              bottom=Side(style='medium'))

        max_row = ws.max_row
        max_col = ws.max_column

        for row in ws[f'B2:{chr(65 + max_col-1)}2']:
            for cell in row:
                cell.border = thick_border

        for row in ws[f'B3:{chr(65 + max_col - 1)}{max_row}']:
            for cell in row:
                cell.border = thin_border

        for col in range(2, max_col + 1):
            ws.cell(row=2, column=col).border = Border(left=ws.cell(row=2, column=col).border.left,
                                                       right=ws.cell(row=2, column=col).border.right,
                                                       top=Side(style='medium'),
                                                       bottom=ws.cell(row=2, column=col).border.bottom)
            ws.cell(row=max_row, column=col).border = Border(left=ws.cell(row=max_row, column=col).border.left,
                                                             right=ws.cell(row=max_row, column=col).border.right,
                                                             top=ws.cell(row=max_row, column=col).border.top,
                                                             bottom=Side(style='medium'))

        for row in range(2, max_row + 1):
            ws.cell(row=row, column=2).border = Border(left=Side(style='medium'),
                                                       right=ws.cell(row=row, column=2).border.right,
                                                       top=ws.cell(row=row, column=2).border.top,
                                                       bottom=ws.cell(row=row, column=2).border.bottom)
            ws.cell(row=row, column=max_col).border = Border(left=ws.cell(row=row, column=max_col).border.left,
                                                             right=Side(style='medium'),
                                                             top=ws.cell(row=row, column=max_col).border.top,
                                                             bottom=ws.cell(row=row, column=max_col).border.bottom)

        for column_cells in ws.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = length + 2

        wb.save(file_path)

        window.show_popup_message2("Table Saved","Results table has been saved to the Excel file.")

    except Exception as e:
        wx.MessageBox(f"Error saving results table: {str(e)}", "Error", wx.OK | wx.ICON_ERROR)


def save_state(window):
    current_sheet = window.sheet_combobox.GetValue()
    state = {
        'Data': deepcopy(window.Data),
        'current_sheet': current_sheet,
        'sheets': {
            sheet: {
                'peak_params_grid': get_grid_data(window.peak_params_grid),
                'peak_count': window.peak_count,
                'selected_peak_index': window.selected_peak_index,
            } for sheet in window.Data['Core levels'].keys()
        },
        'results_grid': get_grid_data(window.results_grid)
    }
    window.history.append(state)
    window.redo_stack.clear()
    if len(window.history) > window.max_history:
        window.history.pop(0)

def undo(window):
    if len(window.history) > 1:
        current_state = window.history.pop()
        window.redo_stack.append(current_state)
        previous_state = window.history[-1]
        restore_state(window, previous_state)

def redo(window):
    if window.redo_stack:
        next_state = window.redo_stack.pop()
        window.history.append(next_state)
        restore_state(window, next_state)


def restore_state(window, state):
    window.Data = deepcopy(state['Data'])

    # Restore sheet-specific data
    for sheet, sheet_data in state['sheets'].items():
        if sheet == state['current_sheet']:
            set_grid_data(window.peak_params_grid, sheet_data['peak_params_grid'])
            window.peak_count = sheet_data['peak_count']
            window.selected_peak_index = sheet_data['selected_peak_index']

    set_grid_data(window.results_grid, state['results_grid'])

    # Switch to the correct sheet
    window.sheet_combobox.SetValue(state['current_sheet'])
    on_sheet_selected(window, state['current_sheet'])

    window.clear_and_replot()

def get_grid_data(grid):
    data = []
    for row in range(grid.GetNumberRows()):
        row_data = [grid.GetCellValue(row, col) for col in range(grid.GetNumberCols())]
        data.append(row_data)
    return data

def set_grid_data(grid, data):
    grid.ClearGrid()
    if len(data) > grid.GetNumberRows():
        grid.AppendRows(len(data) - grid.GetNumberRows())
    for row, row_data in enumerate(data):
        for col, value in enumerate(row_data):
            grid.SetCellValue(row, col, value)