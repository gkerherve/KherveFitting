# libraries/utilities.py
import wx
import numpy as np
import openpyxl
from openpyxl import load_workbook
import os
import pandas as pd
import libraries.Sheet_Operations
import json
from scipy.ndimage import gaussian_filter
from scipy.signal import savgol_filter
from scipy.integrate import cumtrapz


def _clear_peak_params_grid(window):
    num_rows = window.peak_params_grid.GetNumberRows()
    if num_rows > 0:
        window.peak_params_grid.DeleteRows(0, num_rows)
    window.peak_count = 0  # Reset peak count when clearing the grid


def copy_cell(grid):
    print("Copy cell function called")
    print(f"Grid has focus: {grid.HasFocus()}")

    if grid.HasFocus():
        selected_blocks = grid.GetSelectedBlocks()
        if selected_blocks.GetCount() > 0:
            block = selected_blocks.GetBlock(0)
            row, col = block.GetTopRow(), block.GetLeftCol()
            print(f"Selected cell: ({row}, {col})")
            print("CTL C has focus")
            data = grid.GetCellValue(row, col)
            if wx.TheClipboard.Open():
                wx.TheClipboard.SetData(wx.TextDataObject(data))
                wx.TheClipboard.Close()
        else:
            print("No cell selected")
    else:
        print("Grid does not have focus")


def paste_cell(grid):
    if grid.HasFocus():
        selected_blocks = grid.GetSelectedBlocks()
        if selected_blocks.GetCount() > 0 and wx.TheClipboard.Open():
            if wx.TheClipboard.IsSupported(wx.DataFormat(wx.DF_TEXT)):
                data = wx.TextDataObject()
                wx.TheClipboard.GetData(data)
                block = selected_blocks.GetBlock(0)
                row, col = block.GetTopRow(), block.GetLeftCol()
                grid.SetCellValue(row, col, data.GetText())
            wx.TheClipboard.Close()

def load_rsf_data(file_path):
    rsf_dict = {}
    with open(file_path, 'r') as file:
        for line in file:
            parts = line.split()
            if len(parts) == 2:
                core_level, rsf = parts
                rsf_dict[core_level] = float(rsf)
    return rsf_dict


class DraggableText:
    def __init__(self, text):
        self.text = text
        self.press = None
        self.menu = None
        self.connect()

    def connect(self):
        self.cidpress = self.text.figure.canvas.mpl_connect('button_press_event', self.on_press)
        self.cidrelease = self.text.figure.canvas.mpl_connect('button_release_event', self.on_release)
        self.cidmotion = self.text.figure.canvas.mpl_connect('motion_notify_event', self.on_motion)
        self.cidkeypress = self.text.figure.canvas.mpl_connect('key_press_event', self.on_key)
        self.cidrightclick = self.text.figure.canvas.mpl_connect('button_press_event', self.on_right_click)

    def on_right_click(self, event):
        if event.button != 3: return
        if event.inaxes != self.text.axes: return
        contains, _ = self.text.contains(event)
        if not contains: return

        window = event.canvas.GetParent().GetParent().GetParent()
        menu = wx.Menu()

        rotate_item = menu.Append(wx.ID_ANY, "Rotate")
        size_item = menu.Append(wx.ID_ANY, "Change Size")
        delete_item = menu.Append(wx.ID_ANY, "Delete")

        window.Bind(wx.EVT_MENU, self.on_rotate, rotate_item)
        window.Bind(wx.EVT_MENU, self.on_change_size, size_item)
        window.Bind(wx.EVT_MENU, self.on_delete, delete_item)

        window.PopupMenu(menu)
        menu.Destroy()

    def on_rotate(self, event):
        window = self.text.figure.canvas.GetParent().GetParent().GetParent()
        dlg = wx.TextEntryDialog(window, 'Enter rotation angle (degrees):', 'Rotate Text')
        if dlg.ShowModal() == wx.ID_OK:
            try:
                angle = float(dlg.GetValue())
                self.text.set_rotation(angle)
                self.text.figure.canvas.draw()
            except ValueError:
                wx.MessageBox('Please enter a valid number', 'Error')
        dlg.Destroy()

    def on_change_size(self, event):
        window = self.text.figure.canvas.GetParent().GetParent().GetParent()
        dlg = wx.TextEntryDialog(window, 'Enter font size:', 'Change Text Size')
        if dlg.ShowModal() == wx.ID_OK:
            try:
                size = float(dlg.GetValue())
                self.text.set_fontsize(size)
                self.text.figure.canvas.draw()
            except ValueError:
                wx.MessageBox('Please enter a valid number', 'Error')
        dlg.Destroy()

    def on_delete(self, event):
        self.text.remove()
        self.text.figure.canvas.draw()

    def on_press(self, event):
        if event.button != 1: return
        if event.inaxes != self.text.axes: return
        contains, _ = self.text.contains(event)
        if not contains: return
        self.press = self.text.get_position(), event.xdata, event.ydata

    def on_motion(self, event):
        if self.press is None: return
        if event.inaxes != self.text.axes: return
        pos, xpress, ypress = self.press
        dx = event.xdata - xpress
        dy = event.ydata - ypress
        self.text.set_position((pos[0] + dx, pos[1] + dy))
        self.text.figure.canvas.draw()

    def on_release(self, event):
        self.press = None

    def on_key(self, event):
        if event.key == 'delete':
            contains, _ = self.text.contains(event)
            if contains:
                self.text.remove()
                self.text.figure.canvas.draw()


def add_draggable_text(window):
    window.text_mode = not getattr(window, 'text_mode', False)
    if window.text_mode:
        window.canvas.Bind(wx.EVT_LEFT_DOWN, on_canvas_click)
    else:
        window.canvas.Unbind(wx.EVT_LEFT_DOWN)
        window.canvas.mpl_connect('button_press_event', lambda event: None)  # Reset Matplotlib listener


def on_canvas_click(event):
    parent = event.GetEventObject().GetParent()
    while not isinstance(parent, wx.Frame):
        parent = parent.GetParent()
    window = parent

    if not window.text_mode:
        event.Skip()
        return

    x, y = event.GetPosition()
    ax = window.ax
    display_point = ax.transData.inverted().transform((x, y))

    dlg = wx.TextEntryDialog(window, 'Enter text:', 'Add Text Annotation')
    if dlg.ShowModal() == wx.ID_OK:
        text = dlg.GetValue()
        annotation = ax.text(display_point[0], display_point[1], text,
                             picker=5,
                             bbox=dict(facecolor='white', edgecolor='none', alpha=0.7))

        sheet_name = window.sheet_combobox.GetValue()
        if 'Labels' not in window.Data['Core levels'][sheet_name]:
            window.Data['Core levels'][sheet_name]['Labels'] = []

        window.Data['Core levels'][sheet_name]['Labels'].append({
            'text': text,
            'x': display_point[0],
            'y': display_point[1]
        })

        draggable = DraggableText(annotation)
        window.canvas.draw()
    dlg.Destroy()

    window.text_mode = False
    window.canvas.Unbind(wx.EVT_LEFT_DOWN)


def on_delete_sheet(window, event):
    sheet_name = window.sheet_combobox.GetValue()
    dlg = wx.MessageDialog(window, f"Are you sure you want to delete sheet {sheet_name}?",
                           "Confirm Delete", wx.YES_NO | wx.NO_DEFAULT | wx.ICON_QUESTION)

    if dlg.ShowModal() == wx.ID_YES:
        # Remove from Excel file
        wb = openpyxl.load_workbook(window.Data['FilePath'])
        if sheet_name in wb.sheetnames:
            wb.remove(wb[sheet_name])
            wb.save(window.Data['FilePath'])

        # Remove from window.Data
        if sheet_name in window.Data['Core levels']:
            del window.Data['Core levels'][sheet_name]
            window.Data['Number of Core levels'] -= 1

        # Update combobox
        window.sheet_combobox.Delete(window.sheet_combobox.FindString(sheet_name))
        if window.sheet_combobox.GetCount() > 0:
            window.sheet_combobox.SetSelection(0)
            from libraries.Sheet_Operations import on_sheet_selected
            on_sheet_selected(window, window.sheet_combobox.GetString(0))

        # Clear plots if no sheets remain
        if window.sheet_combobox.GetCount() == 0:
            window.ax.clear()
            window.canvas.draw()

    dlg.Destroy()


def rename_sheet(window, new_sheet_name):
    if not new_sheet_name:
        return

    old_sheet_name = window.sheet_combobox.GetValue()
    file_path = window.Data['FilePath']

    # Rename in Excel file
    wb = openpyxl.load_workbook(file_path)
    if old_sheet_name in wb.sheetnames:
        sheet = wb[old_sheet_name]
        sheet.title = new_sheet_name
        wb.save(file_path)

    # Update window.Data
    window.Data['Core levels'][new_sheet_name] = window.Data['Core levels'].pop(old_sheet_name)

    # Update combobox
    index = window.sheet_combobox.FindString(old_sheet_name)
    window.sheet_combobox.Delete(index)
    window.sheet_combobox.Insert(new_sheet_name, index)
    window.sheet_combobox.SetSelection(index)

    # Update plot
    from libraries.Sheet_Operations import on_sheet_selected
    on_sheet_selected(window, new_sheet_name)



class CropWindow(wx.Frame):
    def __init__(self, parent,*args, **kw):
        super().__init__(parent, *args, **kw, style=wx.DEFAULT_FRAME_STYLE & ~(
                    wx.RESIZE_BORDER | wx.MAXIMIZE_BOX | wx.MINIMIZE_BOX | wx.SYSTEM_MENU) | wx.STAY_ON_TOP)
        self.parent = parent
        self.SetTitle("Crop Data")
        self.SetSize(250, 220)

        self.panel = wx.Panel(self)


        sizer = wx.BoxSizer(wx.VERTICAL)

        # Range controls
        range_box = wx.StaticBox(self.panel, label="Crop Range")
        range_sizer = wx.StaticBoxSizer(range_box, wx.VERTICAL)

        min_sizer = wx.BoxSizer(wx.HORIZONTAL)
        self.min_ctrl = wx.SpinCtrlDouble(self.panel, min=0, max=2000, inc=0.1)
        min_sizer.Add(wx.StaticText(self.panel, label="Min BE:"), 0, wx.ALIGN_CENTER_VERTICAL | wx.RIGHT, 5)
        min_sizer.Add(self.min_ctrl, 1)

        max_sizer = wx.BoxSizer(wx.HORIZONTAL)
        self.max_ctrl = wx.SpinCtrlDouble(self.panel, min=0, max=2000, inc=0.1)
        max_sizer.Add(wx.StaticText(self.panel, label="Max BE:"), 0, wx.ALIGN_CENTER_VERTICAL | wx.RIGHT, 5)
        max_sizer.Add(self.max_ctrl, 1)

        range_sizer.Add(min_sizer, 0, wx.EXPAND | wx.ALL, 5)
        range_sizer.Add(max_sizer, 0, wx.EXPAND | wx.ALL, 5)

        # Sheet name control
        name_sizer = wx.BoxSizer(wx.HORIZONTAL)
        self.name_ctrl = wx.TextCtrl(self.panel)
        name_sizer.Add(wx.StaticText(self.panel, label="New Sheet:"), 0, wx.ALIGN_CENTER_VERTICAL | wx.RIGHT, 5)
        name_sizer.Add(self.name_ctrl, 1)

        # Crop button
        self.crop_btn = wx.Button(self.panel, label="Crop")
        self.crop_btn.SetMinSize((125, 40))
        self.crop_btn.Bind(wx.EVT_BUTTON, self.on_crop)

        sizer.Add(range_sizer, 0, wx.EXPAND | wx.ALL, 5)
        sizer.Add(name_sizer, 0, wx.EXPAND | wx.ALL, 5)
        sizer.Add(self.crop_btn, 0, wx.EXPAND | wx.ALL, 5)

        self.panel.SetSizer(sizer)

        self.vline_min = None
        self.vline_max = None

        self.init_values()
        self.Bind(wx.EVT_CLOSE, self.on_close)
        self.min_ctrl.Bind(wx.EVT_SPINCTRLDOUBLE, self.on_range_change)
        self.max_ctrl.Bind(wx.EVT_SPINCTRLDOUBLE, self.on_range_change)

    def init_values(self):
        sheet_name = self.parent.sheet_combobox.GetValue()
        x_values = self.parent.Data['Core levels'][sheet_name]['B.E.']
        min_be = min(x_values) + 2
        max_be = max(x_values) - 2
        self.min_ctrl.SetValue(min_be)
        self.max_ctrl.SetValue(max_be)
        self.name_ctrl.SetValue(f"{sheet_name}_crop")
        self.show_vlines()

    def show_vlines(self):
        if self.vline_min:
            self.vline_min.remove()
        if self.vline_max:
            self.vline_max.remove()

        self.vline_min = self.parent.ax.axvline(self.min_ctrl.GetValue(), color='green', linestyle='--', alpha=0.5)
        self.vline_max = self.parent.ax.axvline(self.max_ctrl.GetValue(), color='green', linestyle='--', alpha=0.5)
        self.parent.canvas.draw_idle()

    def on_range_change(self, event):
        self.show_vlines()

    def on_crop(self, event):
        sheet_name = self.parent.sheet_combobox.GetValue()
        new_name = self.name_ctrl.GetValue()
        min_be = self.min_ctrl.GetValue()
        max_be = self.max_ctrl.GetValue()

        data = self.parent.Data['Core levels'][sheet_name]
        x_values = np.array(data['B.E.'])
        mask = (x_values >= min_be) & (x_values <= max_be)

        # Create new sheet data
        new_data = {
            'B.E.': x_values[mask].tolist(),
            'Raw Data': np.array(data['Raw Data'])[mask].tolist(),
            'Background': {'Bkg Y': np.array(data['Background']['Bkg Y'])[mask].tolist()},
            'Transmission': np.ones(sum(mask)).tolist()
        }

        # Update window.Data
        self.parent.Data['Core levels'][new_name] = new_data
        self.parent.Data['Number of Core levels'] += 1

        # Update Excel file
        wb = openpyxl.load_workbook(self.parent.Data['FilePath'])
        df = pd.DataFrame({
            'Binding Energy': new_data['B.E.'],
            'Raw Data': new_data['Raw Data'],
            'Background': new_data['Background']['Bkg Y'],
            'Transmission': new_data['Transmission']
        })
        # df = pd.DataFrame({
        #     'BE': x,
        #     'Modified Data': y,
        #     'Raw Data': self.parent.Data['Core levels'][original_sheet]['Raw Data'][:len(x)],  # Slice to match x length
        #     'Transmission': np.ones_like(x)
        # })

        with pd.ExcelWriter(self.parent.Data['FilePath'], engine='openpyxl', mode='a') as writer:
            df.to_excel(writer, sheet_name=new_name, index=False)

        # Update JSON file
        json_file_path = os.path.splitext(self.parent.Data['FilePath'])[0] + '.json'
        if os.path.exists(json_file_path):
            from libraries.Save import convert_to_serializable_and_round
            json_data = convert_to_serializable_and_round(self.parent.Data)
            with open(json_file_path, 'w') as json_file:
                json.dump(json_data, json_file, indent=2)

        # Update sheet list
        self.parent.sheet_combobox.Append(new_name)
        self.parent.sheet_combobox.SetValue(new_name)
        from libraries.Sheet_Operations import on_sheet_selected
        on_sheet_selected(self.parent, new_name)

        self.Close()

    def on_close(self, event):
        if self.vline_min:
            self.vline_min.remove()
        if self.vline_max:
            self.vline_max.remove()
        self.parent.canvas.draw_idle()
        self.Destroy()


class PlotModWindow(wx.Frame):
    def __init__(self, parent,*args, **kw):
        super().__init__(parent, *args, **kw, style=wx.DEFAULT_FRAME_STYLE & ~(
                    wx.RESIZE_BORDER | wx.MAXIMIZE_BOX | wx.MINIMIZE_BOX | wx.SYSTEM_MENU) | wx.STAY_ON_TOP)


        self.SetTitle("Plot Modifications")
        self.SetSize(340, 390)

        self.parent = parent
        panel = wx.Panel(self)

        grid_sizer = wx.GridBagSizer(5, 5)

        # First row - Smoothing
        smooth_box = wx.StaticBox(panel, label="Smoothing")
        smooth_sizer = wx.StaticBoxSizer(smooth_box, wx.VERTICAL)

        self.smooth_method = wx.ComboBox(panel, choices=["Gaussian", "Savitzky-Golay", "Moving Average"],
                                         style=wx.CB_READONLY)
        self.smooth_method.SetValue("Gaussian")
        self.smooth_width = wx.SpinCtrl(panel, min=1, max=100, initial=5)

        smooth_sizer.Add(wx.StaticText(panel, label="Method:"), 0, wx.ALL, 5)
        smooth_sizer.Add(self.smooth_method, 0, wx.EXPAND | wx.ALL, 5)
        smooth_sizer.Add(wx.StaticText(panel, label="Width:"), 0, wx.ALL, 5)
        smooth_sizer.Add(self.smooth_width, 0, wx.EXPAND | wx.ALL, 5)

        smooth_btn = wx.Button(panel, label="Apply Smoothing")
        smooth_btn.SetMinSize((125, 40))
        smooth_btn.Bind(wx.EVT_BUTTON, self.on_smooth)
        smooth_sizer.Add(smooth_btn, 0, wx.EXPAND | wx.ALL, 5)

        # First row - Differentiation
        diff_box = wx.StaticBox(panel, label="Differentiation")
        diff_sizer = wx.StaticBoxSizer(diff_box, wx.VERTICAL)

        self.diff_width = wx.SpinCtrl(panel, min=1, max=100, initial=5)
        diff_sizer.Add(wx.StaticText(panel, label="Width:"), 0, wx.ALL, 5)
        diff_sizer.Add(self.diff_width, 0, wx.EXPAND | wx.ALL, 5)

        diff_btn = wx.Button(panel, label="Apply Differentiation")
        diff_btn.SetMinSize((125, 40))
        diff_btn.Bind(wx.EVT_BUTTON, self.on_differentiate)
        diff_sizer.Add(diff_btn, 0, wx.EXPAND | wx.ALL, 5)

        # Second row - Integration
        int_box = wx.StaticBox(panel, label="Integration")
        int_sizer = wx.StaticBoxSizer(int_box, wx.VERTICAL)

        self.int_width = wx.SpinCtrl(panel, min=1, max=100, initial=5)
        int_sizer.Add(wx.StaticText(panel, label="Width:"), 0, wx.ALL, 5)
        int_sizer.Add(self.int_width, 0, wx.EXPAND | wx.ALL, 5)

        int_btn = wx.Button(panel, label="Apply Integration")
        int_btn.SetMinSize((125, 40))
        int_btn.Bind(wx.EVT_BUTTON, self.on_integrate)
        int_sizer.Add(int_btn, 0, wx.EXPAND | wx.ALL, 5)

        # Add to grid
        grid_sizer.Add(smooth_sizer, pos=(0, 0), flag=wx.EXPAND | wx.ALL, border=5)
        grid_sizer.Add(diff_sizer, pos=(0, 1), flag=wx.EXPAND | wx.ALL, border=5)
        grid_sizer.Add(int_sizer, pos=(1, 0), flag=wx.EXPAND | wx.ALL, border=5)

        panel.SetSizer(grid_sizer)
        self.Centre()

    def save_modified_data(self, x, y, sheet_name, operation_type):
        import pandas as pd
        import openpyxl

        original_sheet = sheet_name.split('_')[0]

        # Create DataFrame with required columns
        df = pd.DataFrame({
            'BE': x,
            'Modified Data': y,
            'Raw Data': self.parent.Data['Core levels'][original_sheet]['Raw Data'],
            'Transmission': np.ones_like(x)
        })

        # Load workbook and add new sheet
        wb = openpyxl.load_workbook(self.parent.Data['FilePath'])
        with pd.ExcelWriter(self.parent.Data['FilePath'], engine='openpyxl', mode='a',
                            if_sheet_exists='replace') as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)

        # Update window.Data - convert to list only if needed
        self.parent.Data['Core levels'][sheet_name] = {
            'B.E.': x if isinstance(x, list) else x.tolist(),
            'Raw Data': y if isinstance(y, list) else y.tolist(),
            'Background': {'Bkg Y': np.ones_like(x).tolist()},
            'Transmission': np.ones_like(x).tolist()
        }

        # Update JSON file
        json_file_path = os.path.splitext(self.parent.Data['FilePath'])[0] + '.json'
        if os.path.exists(json_file_path):
            from libraries.Save import convert_to_serializable_and_round
            json_data = convert_to_serializable_and_round(self.parent.Data)
            with open(json_file_path, 'w') as json_file:
                json.dump(json_data, json_file, indent=2)

        # Update sheet list
        self.parent.sheet_combobox.Append(sheet_name)
        self.parent.sheet_combobox.SetValue(sheet_name)
        from libraries.Sheet_Operations import on_sheet_selected
        on_sheet_selected(self.parent, sheet_name)

    def on_smooth(self, event):
        sheet_name = self.parent.sheet_combobox.GetValue()
        method = self.smooth_method.GetValue()
        width = self.smooth_width.GetValue()

        x = self.parent.Data['Core levels'][sheet_name]['B.E.']
        y = self.parent.Data['Core levels'][sheet_name]['Raw Data']

        if method == "Gaussian":
            smoothed = gaussian_filter(y, width)
        elif method == "Savitzky-Golay":
            smoothed = savgol_filter(y, width, 3)
        else:
            kernel = np.ones(width) / width
            smoothed = np.convolve(y, kernel, mode='same')

        new_sheet = f"{sheet_name}_s"
        self.save_modified_data(x, smoothed, new_sheet, "Smoothed")

    def on_differentiate(self, event):
        sheet_name = self.parent.sheet_combobox.GetValue()
        width = self.diff_width.GetValue()

        x = self.parent.Data['Core levels'][sheet_name]['B.E.']
        y = self.parent.Data['Core levels'][sheet_name]['Raw Data']

        derivative = np.gradient(y, x)
        smoothed_deriv = savgol_filter(derivative, width, 3)

        new_sheet = f"{sheet_name}_d"
        self.save_modified_data(x, smoothed_deriv, new_sheet, "Differentiated")

    def on_integrate(self, event):
        sheet_name = self.parent.sheet_combobox.GetValue()
        width = self.int_width.GetValue()

        x = self.parent.Data['Core levels'][sheet_name]['B.E.']
        y = self.parent.Data['Core levels'][sheet_name]['Raw Data']

        integrated = cumtrapz(y, x, initial=0)
        smoothed_int = savgol_filter(integrated, width, 3)

        new_sheet = f"{sheet_name}_i"
        self.save_modified_data(x, smoothed_int, new_sheet, "Integrated")

