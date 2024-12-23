import wx
import os
import json
import openpyxl
import wx
import re
import pandas as pd
import shutil
from vamas import Vamas
from openpyxl import Workbook
from openpyxl.styles import Alignment

from libraries.ConfigFile import Init_Measurement_Data, add_core_level_Data
from libraries.Save import update_undo_redo_state, save_state
from libraries.Sheet_Operations import on_sheet_selected
from libraries.Grid_Operations import populate_results_grid
class ExcelDropTarget(wx.FileDropTarget):
    def __init__(self, window):
        wx.FileDropTarget.__init__(self)
        self.window = window

    def OnDropFiles(self, x, y, filenames):
        from libraries.Open import open_xlsx_file, open_vamas_file
        for file in filenames:
            if file.lower().endswith('.xlsx'):
                wx.CallAfter(open_xlsx_file, self.window, file)
                return True
            elif file.lower().endswith('.vms'):
                wx.CallAfter(open_vamas_file, self.window, file)
                return True
        return False




def load_library_data():
   wb = openpyxl.load_workbook('KherveFitting_library.xlsx')
   sheet = wb['Library']
   data = {}
   for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
       element, orbital, full_name, auger, ke_be, position, ds, rsf, instrument = row
       key = (element, orbital)
       if key not in data:
           data[key] = {}
       data[key][instrument] = {
           'position': position,
           'ds': ds,
           'rsf': rsf,
           'row': row_idx
       }
   return data


def load_library_data_NEWBUTNO():
    wb = openpyxl.load_workbook('KherveFitting_library.xlsx')
    sheet = wb['Library']
    data = {}
    instruments = set()  # Create set for unique instruments

    for row in sheet.iter_rows(min_row=2, values_only=True):
        element, orbital, full_name, auger, ke_be, position, ds, rsf, instrument = row
        instruments.add(instrument)  # Add each instrument to set
        key = (element, orbital)
        if key not in data:
            data[key] = {}
        data[key][instrument] = {
            'position': position,
            'ds': ds,
            'rsf': rsf
        }
    return data, sorted(list(instruments))  # Return data and sorted instruments list

def load_recent_files_from_config(window):
    config = window.load_config()
    window.recent_files = config.get('recent_files', [])
    update_recent_files_menu(window)

def update_recent_files(window, file_path):
    if file_path in window.recent_files:
        window.recent_files.remove(file_path)
    window.recent_files.insert(0, file_path)
    window.recent_files = window.recent_files[:window.max_recent_files]
    update_recent_files_menu(window)
    window.save_config()  # Call save_config directly on the window object

def import_avantage_file(window):
    # Open file dialog to select the Avantage Excel file
    with wx.FileDialog(window, "Open Avantage Excel file", wildcard="Excel files (*.xlsx)|*.xlsx",
                       style=wx.FD_OPEN | wx.FD_FILE_MUST_EXIST) as fileDialog:
        if fileDialog.ShowModal() == wx.ID_CANCEL:
            return

        file_path = fileDialog.GetPath()

    # Load the selected workbook
    wb = openpyxl.load_workbook(file_path)

    # List to keep track of sheets to be removed
    sheets_to_remove = []

    # Iterate through all sheets in the workbook
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]

        if "Survey" in sheet_name or "Scan" in sheet_name:
            # Process Survey or Scan sheets
            if "Survey" in sheet_name or "survey" in sheet_name:
                new_name = "Survey XPS"
            else:
                new_name = sheet_name.split()[0]

            wb.create_sheet(new_name)
            new_sheet = wb[new_name]
            new_sheet['A1'] = "Binding Energy"
            new_sheet['B1'] = "Raw Data"
            # Copy data starting from row 17, skipping column B
            for row in sheet.iter_rows(min_row=17, values_only=True):
                new_sheet.append([row[0]] + list(row[2:]))
            sheets_to_remove.append(sheet_name)

            # Remove data from columns C to X
            for col in new_sheet.iter_cols(min_col=3, max_col=24):
                for cell in col:
                    cell.value = None

        else:
            # Mark other sheets for removal
            sheets_to_remove.append(sheet_name)

    # Remove the original sheets
    for sheet_name in sheets_to_remove:
        del wb[sheet_name]

    # Save the modified workbook with a new name
    new_file_path = os.path.splitext(file_path)[0] + "_Kfitting.xlsx"
    wb.save(new_file_path)

    # Open the newly created file using the existing open_xlsx_file function
    from libraries.Open import open_xlsx_file
    open_xlsx_file(window, new_file_path)


def parse_avg_file(file_path):
    with open(file_path, 'r') as file:
        content = file.read()

    photon_energy = float(re.search(r'DS_SOPROPID_ENERGY\s+:\s+VT_R4\s+=\s+(\d+\.\d+)', content).group(1))
    start_energy, width, num_points = map(float, re.search(r'\$SPACEAXES=1\s+0=\s+(\d+\.\d+),\s+(\d+\.\d+),\s+(\d+),',
                                                           content).groups())

    # Modified part to handle multiple numbers per line
    y_values = []
    for match in re.findall(r'LIST@\s+\d+=\s+([\d., ]+)', content):
        values = [float(val.strip()) for val in match.split(',')]
        y_values.extend(values)

    return photon_energy, start_energy, width, int(num_points), y_values


def create_excel_from_avg(avg_file_path):
    sheet_name = os.path.basename(avg_file_path).split()[0]
    photon_energy, start_energy, width, num_points, y_values = parse_avg_file(avg_file_path)

    be_values = [photon_energy - (start_energy + i * width) for i in range(num_points)]

    df = pd.DataFrame({
        'BE': be_values,
        'Intensity': y_values[:num_points]  # Use only the correct number of y values
    })

    output_path = avg_file_path.rsplit('.', 1)[0] + '.xlsx'
    with pd.ExcelWriter(output_path) as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)

    return output_path


def open_avg_file(window):
    with wx.FileDialog(window, "Open AVG file", wildcard="AVG files (*.avg)|*.avg",
                       style=wx.FD_OPEN | wx.FD_FILE_MUST_EXIST) as fileDialog:
        if fileDialog.ShowModal() == wx.ID_CANCEL:
            return

        avg_file_path = fileDialog.GetPath()

    try:
        excel_file_path = create_excel_from_avg(avg_file_path)

        from libraries.Open import open_xlsx_file
        open_xlsx_file(window, excel_file_path)
    except Exception as e:
        wx.MessageBox(f"Error processing AVG file: {str(e)}", "Error", wx.OK | wx.ICON_ERROR)


def import_multiple_avg_files(window):
    with wx.DirDialog(window, "Choose a directory containing AVG files",
                      style=wx.DD_DEFAULT_STYLE | wx.DD_DIR_MUST_EXIST) as dirDialog:

        if dirDialog.ShowModal() == wx.ID_CANCEL:
            return

        folder_path = dirDialog.GetPath()

    try:
        avg_files = [f for f in os.listdir(folder_path) if f.lower().endswith('.avg')]

        if not avg_files:
            wx.MessageBox("No AVG files found in the selected folder.", "Information", wx.OK | wx.ICON_INFORMATION)
            return

        folder_name = os.path.basename(folder_path)
        excel_file_path = os.path.join(folder_path, f"{folder_name}.xlsx")

        with pd.ExcelWriter(excel_file_path) as writer:
            for avg_file in avg_files:
                avg_file_path = os.path.join(folder_path, avg_file)
                sheet_name = os.path.splitext(avg_file)[0]  # Use file name without extension as sheet name

                photon_energy, start_energy, width, num_points, y_values = parse_avg_file(avg_file_path)
                be_values = [photon_energy - (start_energy + i * width) for i in range(num_points)]

                df = pd.DataFrame({
                    'BE': be_values,
                    'Intensity': y_values[:num_points]
                })

                df.to_excel(writer, sheet_name=sheet_name, index=False)

        wx.MessageBox(f"Excel file created: {excel_file_path}", "Success", wx.OK | wx.ICON_INFORMATION)

        # Open the created Excel file
        from libraries.Open import open_xlsx_file
        open_xlsx_file(window, excel_file_path)

    except Exception as e:
        wx.MessageBox(f"Error processing AVG files: {str(e)}", "Error", wx.OK | wx.ICON_ERROR)


def open_xlsx_file(window, file_path=None):
    """
    Opens an Excel file, loads its data, and updates the application's state accordingly.
    If a corresponding JSON file exists, it loads data from there instead.
    """
    print("Starting open_xlsx_file function")
    if file_path is None:
        with wx.FileDialog(window, "Open XLSX file", wildcard="Excel files (*.xlsx)|*.xlsx",
                           style=wx.FD_OPEN | wx.FD_FILE_MUST_EXIST) as dlg:
            if dlg.ShowModal() == wx.ID_OK:
                file_path = dlg.GetPath()
            else:
                return

    window.SetStatusText(f"Selected File: {file_path}", 0)

    try:
        # Clear undo and redo history
        window.history = []
        window.redo_stack = []
        update_undo_redo_state(window)

        # Clear the results grid
        window.results_grid.ClearGrid()
        if window.results_grid.GetNumberRows() > 0:
            window.results_grid.DeleteRows(0, window.results_grid.GetNumberRows())

        # Look for corresponding .json file
        json_file = os.path.splitext(file_path)[0] + '.json'
        if os.path.exists(json_file):
            print(f"Found corresponding .json file: {json_file}")
            with open(json_file, 'r') as f:
                loaded_data = json.load(f)

            # Convert data structure without changing types
            window.Data = convert_from_serializable(loaded_data)

            print("Loaded data from .json file")

            # Populate the results grid
            populate_results_grid(window)
        else:
            print("No corresponding .json file found. Initializing new data.")
            # Initialize the measurement data
            window.Data = Init_Measurement_Data(window)

        # Read the Excel file
        excel_file = pd.ExcelFile(file_path)
        all_sheet_names = excel_file.sheet_names
        sheet_names = [name for name in all_sheet_names if
                       name.lower() not in ["results table", "experimental description"]]

        results_table_index = -1
        for i, name in enumerate(all_sheet_names):
            if name.lower() == "results table":
                results_table_index = i
                break

        if results_table_index != -1:
            sheet_names = sheet_names[:results_table_index]

        print(f"Number of sheets: {len(sheet_names)}")

        # Update file path
        window.Data['FilePath'] = file_path

        # If we didn't load from json, populate the data from Excel
        if 'Core levels' not in window.Data or not window.Data['Core levels']:
            window.Data['Number of Core levels'] = 0
            for sheet_name in sheet_names:
                window.Data = add_core_level_Data(window.Data, window, file_path, sheet_name)

        print(f"Final number of core levels: {window.Data['Number of Core levels']}")

        # Load BE correction
        window.load_be_correction()

        # Update sheet names in the combobox
        window.sheet_combobox.Clear()
        window.sheet_combobox.AppendItems(sheet_names)

        # Set the first sheet as the selected one
        first_sheet = sheet_names[0]
        window.sheet_combobox.SetValue(first_sheet)

        # Use on_sheet_selected to update peak parameter grid and plot
        event = wx.CommandEvent(wx.EVT_COMBOBOX.typeId)
        event.SetString(first_sheet)
        window.plot_config.plot_limits.clear()
        on_sheet_selected(window, event)

        # undo and redo
        save_state(window)

        # Update recent files list
        update_recent_files(window, file_path)

        print("open_xlsx_file function completed successfully")
    except Exception as e:
        print(f"Error in open_xlsx_file: {str(e)}")
        import traceback
        traceback.print_exc()
        wx.MessageBox(f"Error reading file: {str(e)}", "Error", wx.OK | wx.ICON_ERROR)

def convert_from_serializable(obj):
    """
    Recursively converts a serializable object (list or dict) back into its original structure.
    This is used to restore complex data structures that were serialized to JSON.
    """
    if isinstance(obj, list):
        return [convert_from_serializable(item) for item in obj]
    elif isinstance(obj, dict):
        return {k: convert_from_serializable(v) for k, v in obj.items()}
    else:
        return obj

def update_recent_files(window, file_path):
    """
    Updates the list of recently opened files, adding the new file to the top and removing duplicates.
    """
    if file_path in window.recent_files:
        window.recent_files.remove(file_path)
    window.recent_files.insert(0, file_path)
    window.recent_files = window.recent_files[:window.max_recent_files]
    update_recent_files_menu(window)
    save_recent_files_to_config(window)

def update_recent_files_menu(window):
    """
    Refreshes the 'Recent Files' menu with the updated list of recently opened files.
    """
    if hasattr(window, 'recent_files_menu'):
        # Remove all existing menu items
        for item in window.recent_files_menu.GetMenuItems():
            window.recent_files_menu.DestroyItem(item)

        # Add new menu items for recent files
        for i, file_path in enumerate(window.recent_files):
            item = window.recent_files_menu.Append(wx.ID_ANY, os.path.basename(file_path))
            window.Bind(wx.EVT_MENU, lambda evt, fp=file_path: open_xlsx_file(window, fp), item)

def save_recent_files_to_config(window):
    """
    Saves the updated list of recently opened files to the application's configuration.
    """
    window.recent_files = window.recent_files[:window.max_recent_files]
    window.save_config()


def open_vamas_file_dialog(window):
    """
    Open a file dialog for selecting a VAMAS file and process it.

    This function displays a file dialog for the user to select a VAMAS file,
    then calls open_vamas_file to process the selected file.

    Args:
    window: The main application window object.
    """
    with wx.FileDialog(window, "Open VAMAS file", wildcard="VAMAS files (*.vms)|*.vms",
                       style=wx.FD_OPEN | wx.FD_FILE_MUST_EXIST) as fileDialog:
        if fileDialog.ShowModal() == wx.ID_CANCEL:
            return
        file_path = fileDialog.GetPath()
        open_vamas_file(window, file_path)

def open_vamas_file(window, file_path):
    """
    Open and process a VAMAS file, converting it to an Excel file format.
    This function reads a VAMAS file, extracts its data and metadata,
    and creates a new Excel file with multiple sheets for each data block
    and an additional sheet for experimental description.

    Args:
    window: The main application window object.
    file_path: The path to the VAMAS file to be opened.
    """
    try:
        # Clear undo and redo history
        window.history = []
        window.redo_stack = []
        update_undo_redo_state(window)

        if not os.path.exists(file_path):
            raise FileNotFoundError(f"The file {file_path} does not exist.")

        # Copy VAMAS file to current working directory
        vamas_filename = os.path.basename(file_path)
        destination_path = os.path.join(os.getcwd(), vamas_filename)
        shutil.copy2(file_path, destination_path)

        # Read VAMAS file
        vamas_data = Vamas(vamas_filename)

        # Create new Excel workbook
        wb = Workbook()
        wb.remove(wb.active)

        exp_data = []  # Store experimental description data

        # Process each block in the VAMAS file
        for i, block in enumerate(vamas_data.blocks, start=1):
            # Determine sheet name
            if block.species_label.lower() == "wide" or block.transition_or_charge_state_label.lower() == "none":
                sheet_name = block.species_label
            else:
                sheet_name = f"{block.species_label}{block.transition_or_charge_state_label}"
            sheet_name = sheet_name.replace("/", "_")

            # Create new sheet
            ws = wb.create_sheet(title=sheet_name)

            # Extract and process data
            num_points = block.num_y_values
            x_start = block.x_start
            x_step = block.x_step
            x_values = [x_start + i * x_step for i in range(num_points)]
            y_values = block.corresponding_variables[0].y_values
            y_unit = block.corresponding_variables[0].unit
            num_scans = block.num_scans_to_compile_block

            # Convert counts to counts per second if necessary
            if y_unit != "c/s":
                y_values = [y / num_scans for y in y_values]

            # Convert to Binding Energy if necessary
            if block.x_label == "Kinetic Energy":
                x_values = [window.photons - x - window.workfunction for x in x_values]
                x_label = "Binding Energy"
            else:
                x_label = block.x_label

            # Write data to sheet
            ws.append([x_label, "Intensity"])
            for x, y in zip(x_values, y_values):
                ws.append([x, y])

            # Store experimental setup data
            exp_data.append([
                f"Block {i}",
                block.sample_identifier,
                f"{block.year}/{block.month}/{block.day}",
                f"{block.hour}:{block.minute}:{block.second}",
                block.technique,
                f"{block.species_label} {block.transition_or_charge_state_label}",
                block.num_scans_to_compile_block,
                block.analysis_source_label,
                block.analysis_source_characteristic_energy,
                block.analysis_source_beam_width_x,
                block.analysis_source_beam_width_y,
                block.analyzer_pass_energy_or_retard_ratio_or_mass_res,
                block.analyzer_work_function_or_acceptance_energy,
                block.analyzer_mode,
                block.sputtering_source_energy if hasattr(block, 'sputtering_source_energy') else 'N/A',
                block.analyzer_axis_take_off_polar_angle,
                block.analyzer_axis_take_off_azimuth,
                block.target_bias,
                block.analysis_width_x,
                block.analysis_width_y,
                block.x_label,
                block.x_units,
                block.x_start,
                block.x_step,
                block.num_y_values,
                block.num_scans_to_compile_block,
                block.signal_collection_time,
                block.signal_time_correction,
                y_unit,
                block.num_lines_block_comment,
                block.block_comment
            ])

        # Create "Experimental description" sheet
        exp_sheet = wb.create_sheet(title="Experimental description")
        exp_sheet.column_dimensions['A'].width = 50
        exp_sheet.column_dimensions['B'].width = 100
        left_aligned = Alignment(horizontal='left')

        # Add VAMAS header information
        exp_sheet.append(["VAMAS Header Information"])
        for item in [
            ("Format Identifier", vamas_data.header.format_identifier),
            ("Institution Identifier", vamas_data.header.institution_identifier),
            ("Instrument Model", vamas_data.header.instrument_model_identifier),
            ("Operator Identifier", vamas_data.header.operator_identifier),
            ("Experiment Identifier", vamas_data.header.experiment_identifier),
            ("Number of Comment Lines", vamas_data.header.num_lines_comment),
            ("Comment", vamas_data.header.comment),
            ("Experiment Mode", vamas_data.header.experiment_mode),
            ("Scan Mode", vamas_data.header.scan_mode),
            ("Number of Spectral Regions", vamas_data.header.num_spectral_regions),
            ("Number of Analysis Positions", vamas_data.header.num_analysis_positions),
            ("Number of Discrete X Coordinates", vamas_data.header.num_discrete_x_coords_in_full_map),
            ("Number of Discrete Y Coordinates", vamas_data.header.num_discrete_y_coords_in_full_map)
        ]:
            exp_sheet.append(item)

        exp_sheet.append([])  # Add a blank row for separation

        # Define the order of block information
        block_info_order = [
            "Sample ID", "Year/Month/Day", "Time HH,MM,SS", "Technique", "Species & Transition", "Number of scans",
            "Source Label", "Source Energy", "Source width X", "Source width Y", "Pass Energy", "Work Function",
            "Analyzer Mode", "Sputtering Energy", "Take-off Polar Angle", "Take-off Azimuth", "Target Bias",
            "Analysis Width X", "Analysis Width Y", "X Label", "X Units", "X Start", "X Step", "Num Y Values",
            "Num Scans", "Collection Time", "Time Correction", "Y Unit", "# Comment Lines", "Block Comment"
        ]

        # Add block information
        for i, block_data in enumerate(exp_data, start=1):
            exp_sheet.append([f"Block {i}", ""])
            for j, info in enumerate(block_info_order):
                exp_sheet.append([info, block_data[j + 1]])
            exp_sheet.append([])  # Add a blank row between blocks

        # Set alignment for all cells in column B
        for row in exp_sheet.iter_rows(min_row=1, max_row=exp_sheet.max_row, min_col=2, max_col=2):
            for cell in row:
                cell.alignment = left_aligned

        # Save Excel file
        excel_filename = os.path.splitext(vamas_filename)[0] + ".xlsx"
        excel_path = os.path.join(os.path.dirname(file_path), excel_filename)
        wb.save(excel_path)

        # Remove temporary VAMAS file
        os.remove(destination_path)

        # Update window.Data with the new Excel file
        window.Data = Init_Measurement_Data(window)
        window.Data['FilePath'] = excel_path

        # Open the Excel file and populate window.Data
        open_xlsx_file_vamas(window, excel_path)

    except FileNotFoundError as e:
        wx.MessageBox(f"File not found: {str(e)}", "Error", wx.OK | wx.ICON_ERROR)
    except Exception as e:
        wx.MessageBox(f"Error processing VAMAS file: {str(e)}", "Error", wx.OK | wx.ICON_ERROR)

def open_xlsx_file_vamas(window, file_path):
    """
    Open and process an Excel file created from a VAMAS file.

    This function initializes the data structure, reads the Excel file,
    populates the window.Data dictionary with core level information,
    updates the GUI elements, and plots the data for the first sheet.

    Args:
    window: The main application window object.
    file_path: The path to the Excel file to be opened.
    """
    try:
        # Update status bar with the selected file path
        window.SetStatusText(f"Selected File: {file_path}", 0)

        # Initialize the measurement data structure
        window.Data = Init_Measurement_Data(window)
        window.Data['FilePath'] = file_path

        # Read the Excel file
        excel_file = pd.ExcelFile(file_path)

        # Get sheet names, excluding the "Experimental description" sheet
        sheet_names = [name for name in excel_file.sheet_names if name != "Experimental description"]

        # Initialize the number of core levels
        window.Data['Number of Core levels'] = 0

        # Process each sheet (core level) in the Excel file
        for sheet_name in sheet_names:
            window.Data = add_core_level_Data(window.Data, window, file_path, sheet_name)

        print(f"Final number of core levels: {window.Data['Number of Core levels']}")

        # Update GUI elements
        window.sheet_combobox.Clear()
        window.sheet_combobox.AppendItems(sheet_names)
        window.sheet_combobox.SetValue(sheet_names[0])  # Set first sheet as default

        # Plot the data for the first sheet
        window.plot_manager.plot_data(window)

    except Exception as e:
        # Handle any errors that occur during file processing
        print(f"Error in open_xlsx_file_vamas: {str(e)}")
        import traceback
        traceback.print_exc()
        wx.MessageBox(f"Error reading Excel file: {str(e)}", "Error", wx.OK | wx.ICON_ERROR)



# ------------------ HISTORRY DEF ---------------------------------------------------
# -----------------------------------------------------------------------------------
